import os
from datetime import datetime
import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Global list to collect details from each test case
run_details = []

@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    """
    Hook to capture test execution status, duration, and errors.
    """
    outcome = yield
    rep = outcome.get_result()
    
    # Capture results during the actual 'call' phase of a test
    if rep.when == "call" or (rep.when == "setup" and rep.failed):
        suite_name = item.parent.name if item.parent else "Landing Page Suite"
        # Use docstring as description if available, otherwise fallback to function name
        step_name = item.obj.__doc__.strip() if item.obj.__doc__ else item.name
        # Clean whitespaces
        step_name = " ".join(step_name.split())
        
        status = "PASS" if rep.passed else "FAIL"
        duration = int(rep.duration * 1000)  # Convert seconds to milliseconds
        error = str(rep.longrepr) if rep.failed else "N/A"
        
        run_details.append({
            "timestamp": datetime.now().strftime("%I:%M:%S %p"),
            "suiteName": suite_name,
            "stepName": step_name,
            "status": status,
            "duration": f"{duration}ms",
            "error": error
        })

def pytest_sessionfinish(session, exitstatus):
    """
    pytest lifecycle finish hook: Compiles collected test results
    and outputs them to a formatted Excel file.
    """
    # Ensure output directory (reports/) exists in the project root
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    reports_dir = os.path.join(base_dir, "reports")
    os.makedirs(reports_dir, exist_ok=True)
    report_path = os.path.join(reports_dir, "E2E_Web_Test_Report.xlsx")
    
    wb = Workbook()
    
    # ── 1. Summary Sheet ──────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_summary.merge_cells("A1:C2")
    title_cell = ws_summary["A1"]
    title_cell.value = "Smart Press Web E2E Automation Test Summary"
    title_cell.font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(vertical="center", horizontal="center")
    # Teal color matching original design (fgColor: 0F766E)
    title_cell.fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    
    # Metric Calculations
    total_steps = len(run_details)
    passed_steps = sum(1 for step in run_details if step["status"] == "PASS")
    failed_steps = total_steps - passed_steps
    pass_rate = f"{((passed_steps / total_steps) * 100):.1f}%" if total_steps > 0 else "0%"
    
    ws_summary.append([])  # Blank row 3
    ws_summary.append([])  # Blank row 4
    
    # Data Table Header
    ws_summary.append(["Metric / Property", "Value"])  # Row 5
    header_row = ws_summary[5]
    for cell in header_row:
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        
    # Stats Data Rows
    stats_data = [
        ("Date of Execution", datetime.now().strftime("%m/%d/%Y %I:%M:%S %p")),
        ("Total Steps Executed", total_steps),
        ("Steps Passed", passed_steps),
        ("Steps Failed", failed_steps),
        ("Overall Pass Rate", pass_rate)
    ]
    
    for metric, val in stats_data:
        ws_summary.append([metric, val])
        row_idx = ws_summary.max_row
        ws_summary.cell(row=row_idx, column=1).font = Font(name="Segoe UI", bold=True)
        val_cell = ws_summary.cell(row=row_idx, column=2)
        val_cell.font = Font(name="Segoe UI")
        
        # Color conditional styling for Overall Pass Rate
        if metric == "Overall Pass Rate":
            color = "991B1B" if failed_steps > 0 else "065F46"
            val_cell.font = Font(name="Segoe UI", bold=True, color=color)
            
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 30
    
    # ── 2. Details Sheet ──────────────────────────
    ws_details = wb.create_sheet("Execution Details")
    ws_details.views.sheetView[0].showGridLines = True
    
    # Header definitions
    headers = ["Time", "Test Suite", "Test Step Description", "Status", "Duration", "Error / Failure Details"]
    ws_details.append(headers)
    
    detail_header = ws_details[1]
    for cell in detail_header:
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        
    # Populate Details Data
    for step in run_details:
        ws_details.append([
            step["timestamp"],
            step["suiteName"],
            step["stepName"],
            step["status"],
            step["duration"],
            step["error"]
        ])
        row_idx = ws_details.max_row
        
        for col_idx in range(1, 7):
            cell = ws_details.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Segoe UI", size=10)
            
        # Format Status Cell (Color Badges)
        status_cell = ws_details.cell(row=row_idx, column=4)
        status_cell.alignment = Alignment(horizontal="center")
        if step["status"] == "PASS":
            status_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            status_cell.font = Font(name="Segoe UI", size=10, bold=True, color="065F46")
        else:
            status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            status_cell.font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
            
    ws_details.column_dimensions["A"].width = 15
    ws_details.column_dimensions["B"].width = 25
    ws_details.column_dimensions["C"].width = 45
    ws_details.column_dimensions["D"].width = 12
    ws_details.column_dimensions["E"].width = 15
    ws_details.column_dimensions["F"].width = 50
    
    wb.save(report_path)
    print(f"Excel E2E Report saved successfully: {report_path}")
