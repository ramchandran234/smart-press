import os
import time
import threading
from datetime import datetime
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

TARGET_URL = "https://smart-press-backend.onrender.com/"
CONCURRENT_USERS = 100
# Allow environment variable override for local fast validation
DURATION_SECONDS = int(os.environ.get("LOAD_TEST_DURATION", 60))

results = []
stop_event = threading.Event()

def worker():
    session = requests.Session()
    while not stop_event.is_set():
        start_time = time.perf_counter()
        timestamp = datetime.now().strftime("%I:%M:%S %p")
        try:
            # Query the root healthcheck endpoint
            response = session.get(TARGET_URL, timeout=10)
            latency = int((time.perf_counter() - start_time) * 1000)
            status = "SUCCESS" if response.status_code == 200 else "FAILED"
            error_details = "N/A" if response.status_code == 200 else f"HTTP {response.status_code}"
        except Exception as e:
            latency = int((time.perf_counter() - start_time) * 1000)
            status = "FAILED"
            error_details = str(e)
        
        results.append({
            "timestamp": timestamp,
            "latency": latency,
            "status": status,
            "error": error_details
        })

def run_load_test():
    print(f"[START] Starting baseline load test on: {TARGET_URL}")
    print(f"[USERS] Virtual Users: {CONCURRENT_USERS}")
    print(f"[DURATION] Duration: {DURATION_SECONDS} seconds")
    
    threads = []
    for _ in range(CONCURRENT_USERS):
        t = threading.Thread(target=worker)
        t.start()
        threads.append(t)
        
    time.sleep(DURATION_SECONDS)
    stop_event.set()
    
    for t in threads:
        t.join()
        
    total_requests = len(results)
    print(f"[FINISH] Load test completed. Total requests sent: {total_requests}")
    
    if total_requests == 0:
        print("[ERROR] Error: No requests were sent.")
        return
        
    # Process Metrics
    successful_requests = sum(1 for r in results if r["status"] == "SUCCESS")
    failed_requests = total_requests - successful_requests
    avg_latency = int(sum(r["latency"] for r in results) / total_requests)
    min_latency = min(r["latency"] for r in results)
    max_latency = max(r["latency"] for r in results)
    rps = round(total_requests / DURATION_SECONDS, 2)
    success_rate = f"{((successful_requests / total_requests) * 100):.1f}%"
    
    print("\n--- RESULTS SUMMARY ---")
    print(f"RPS: {rps} req/sec")
    print(f"Response Times - Avg: {avg_latency}ms | Min: {min_latency}ms | Max: {max_latency}ms")
    print(f"Success Rate: {success_rate} ({successful_requests}/{total_requests})")
    
    # Save Report
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    reports_dir = os.path.join(base_dir, "reports")
    os.makedirs(reports_dir, exist_ok=True)
    report_path = os.path.join(reports_dir, "Load_Test_Report.xlsx")
    
    wb = Workbook()
    
    # ── 1. Summary Sheet ──────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_summary.merge_cells("A1:C2")
    title_cell = ws_summary["A1"]
    title_cell.value = "Smart Press Web API Load Test Summary"
    title_cell.font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(vertical="center", horizontal="center")
    # Teal color matching existing E2E report
    title_cell.fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    
    ws_summary.append([])  # Blank row 3
    ws_summary.append([])  # Blank row 4
    
    # Data Table Header
    ws_summary.append(["Metric / Property", "Value"])  # Row 5
    header_row = ws_summary[5]
    for cell in header_row:
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        
    stats_data = [
        ("Date of Execution", datetime.now().strftime("%m/%d/%Y %I:%M:%S %p")),
        ("Target API Endpoint", TARGET_URL),
        ("Test Duration", f"{DURATION_SECONDS} seconds"),
        ("Concurrent Virtual Users", CONCURRENT_USERS),
        ("Total Requests Sent", total_requests),
        ("Requests Per Second (RPS)", f"{rps} req/sec"),
        ("Average Response Time", f"{avg_latency} ms"),
        ("Minimum Response Time", f"{min_latency} ms"),
        ("Maximum Response Time", f"{max_latency} ms"),
        ("Successful Requests", successful_requests),
        ("Failed Requests", failed_requests),
        ("Overall Success Rate", success_rate)
    ]
    
    for metric, val in stats_data:
        ws_summary.append([metric, val])
        row_idx = ws_summary.max_row
        ws_summary.cell(row=row_idx, column=1).font = Font(name="Segoe UI", bold=True)
        val_cell = ws_summary.cell(row=row_idx, column=2)
        val_cell.font = Font(name="Segoe UI")
        
        # Color conditional styling for Overall Success Rate
        if metric == "Overall Success Rate":
            color = "991B1B" if failed_requests > 0 else "065F46"
            val_cell.font = Font(name="Segoe UI", bold=True, color=color)
            
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 45
    
    # ── 2. Details Sheet ──────────────────────────
    ws_details = wb.create_sheet("Request Latencies")
    ws_details.views.sheetView[0].showGridLines = True
    
    # Header definitions
    headers = ["Time", "Request Index", "Status", "Latency (ms)", "Error / Failure Details"]
    ws_details.append(headers)
    
    detail_header = ws_details[1]
    for cell in detail_header:
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        
    # Populate Details Data
    for idx, step in enumerate(results, 1):
        ws_details.append([
            step["timestamp"],
            f"Request #{idx}",
            step["status"],
            f"{step['latency']}ms",
            step["error"]
        ])
        row_idx = ws_details.max_row
        
        for col_idx in range(1, 6):
            cell = ws_details.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Segoe UI", size=10)
            
        # Format Status Cell (Color Badges)
        status_cell = ws_details.cell(row=row_idx, column=3)
        status_cell.alignment = Alignment(horizontal="center")
        if step["status"] == "SUCCESS":
            status_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            status_cell.font = Font(name="Segoe UI", size=10, bold=True, color="065F46")
        else:
            status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            status_cell.font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
            
    ws_details.column_dimensions["A"].width = 15
    ws_details.column_dimensions["B"].width = 18
    ws_details.column_dimensions["C"].width = 12
    ws_details.column_dimensions["D"].width = 15
    ws_details.column_dimensions["E"].width = 40
    
    wb.save(report_path)
    print(f"[REPORT] Load Test Excel Report saved successfully: {report_path}")

if __name__ == "__main__":
    run_load_test()
